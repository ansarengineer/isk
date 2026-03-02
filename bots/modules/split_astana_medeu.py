import os
import re
from datetime import datetime
from copy import copy
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

from openpyxl import load_workbook, Workbook


# --- НАСТРОЙКИ: ваши 4 истца и 2 суда, и итоговые имена файлов ---
TARGETS = [
    ("Жагыпарова Гүлжан Нурланқызы", "Медеуский районный суд города Алматы", "Жагыпарова объед Медеу.xlsx"),
    ("Жагыпарова Гүлжан Нурланқызы", "Межрайонный суд по гражданским делам г. Астаны", "Жагыпарова объед Астана.xlsx"),

    ("Каипов Ришат Равильевич", "Медеуский районный суд города Алматы", "Каипов объед Медеу.xlsx"),
    ("Каипов Ришат Равильевич", "Межрайонный суд по гражданским делам г. Астаны", "Каипов объед Астана.xlsx"),

    ("Усманова Асель Сыдыковна", "Медеуский районный суд города Алматы", "Усманова объед Медеу.xlsx"),
    ("Усманова Асель Сыдыковна", "Межрайонный суд по гражданским делам g. Астаны", "Усманова объед Астана.xlsx"),  # на всякий — иногда "г." бывает латиницей
    ("Усманова Асель Сыдыковна", "Межрайонный суд по гражданским делам г. Астаны", "Усманова объед Астана.xlsx"),

    ("ТОО «Коллекторское агентство «Эксперт Плюс»", "Медеуский районный суд города Алматы", "ЭП объед Медеу.xlsx"),
    ("ТОО «Коллекторское агентство «Эксперт Плюс»", "Межрайонный суд по гражданским делам г. Астаны", "ЭП объед Астана.xlsx"),
]

# Чтобы не создавать дубль "Усманова ... Астана.xlsx" из-за 2 строк выше:
def dedupe_targets(targets):
    seen = set()
    out = []
    for p, c, fn in targets:
        key = (p, c, fn)
        if key not in seen:
            seen.add(key)
            out.append((p, c, fn))
    return out

TARGETS = dedupe_targets(TARGETS)


def norm_text(s: str) -> str:
    """Нормализация строк для сравнения: пробелы, регистр, 'г.Астаны' vs 'г. Астаны' и т.п."""
    if s is None:
        return ""
    s = str(s)
    s = s.replace("\u00A0", " ")  # неразрывные пробелы
    s = s.strip().lower()
    s = re.sub(r"\s+", " ", s)
    # унифицируем "г.астаны" варианты
    s = s.replace("г.астаны", "г. астаны")
    s = s.replace("г.алматы", "г. алматы")
    s = s.replace("г. астаны", "г. астаны")
    return s


def copy_sheet_filtered(src_ws, dst_ws, keep_row_indices):
    """
    Копирует строку заголовков (1) и строки по индексам из src_ws в dst_ws,
    пытаясь сохранить стили/ширины.
    """
    # Копируем ширины колонок
    for col_letter, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col_letter].width = dim.width

    # Копируем высоты строк (минимально полезно)
    for r, dim in src_ws.row_dimensions.items():
        if dim.height is not None:
            dst_ws.row_dimensions[r].height = dim.height

    # Какие строки переносим: заголовок + найденные строки
    rows_to_copy = [1] + keep_row_indices

    out_row = 1
    for r in rows_to_copy:
        for c in range(1, src_ws.max_column + 1):
            src_cell = src_ws.cell(row=r, column=c)
            dst_cell = dst_ws.cell(row=out_row, column=c, value=src_cell.value)

            # Стили (если есть)
            if src_cell.has_style:
                dst_cell._style = copy(src_cell._style)
                dst_cell.font = copy(src_cell.font)
                dst_cell.border = copy(src_cell.border)
                dst_cell.fill = copy(src_cell.fill)
                dst_cell.number_format = src_cell.number_format
                dst_cell.protection = copy(src_cell.protection)
                dst_cell.alignment = copy(src_cell.alignment)

        out_row += 1


def find_header_column_indices(ws):
    """Считывает заголовки из 1-й строки: возвращает список названий колонок и маппинг имя->индекс."""
    headers = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        name = "" if v is None else str(v).strip()
        headers.append(name)
    mapping = {h: i + 1 for i, h in enumerate(headers) if h != ""}
    return headers, mapping


def split_excel(input_path, sheet_name, plaintiff_col_name, court_col_name):
    wb = load_workbook(input_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Лист '{sheet_name}' не найден.")
    ws = wb[sheet_name]

    headers, header_map = find_header_column_indices(ws)
    if plaintiff_col_name not in header_map:
        raise ValueError("Колонка истца не найдена в заголовках.")
    if court_col_name not in header_map:
        raise ValueError("Колонка суда не найдена в заголовках.")

    plaintiff_col = header_map[plaintiff_col_name]
    court_col = header_map[court_col_name]

    # Подготовим выходную папку
    base_dir = os.path.dirname(input_path)
    base_name = os.path.splitext(os.path.basename(input_path))[0]
    out_dir = os.path.join(
        base_dir,
        f"{base_name}_split_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    )
    os.makedirs(out_dir, exist_ok=True)

    # Предварительно нормализуем цели
    targets_norm = [(norm_text(p), norm_text(c), fn) for p, c, fn in TARGETS]

    # Собираем индексы строк для каждой цели
    keep_map = {fn: [] for _, _, fn in targets_norm}

    # Пробегаем строки данных (со 2-й строки)
    for r in range(2, ws.max_row + 1):
        p_val = norm_text(ws.cell(row=r, column=plaintiff_col).value)
        c_val = norm_text(ws.cell(row=r, column=court_col).value)

        for tp, tc, fn in targets_norm:
            if p_val == tp and c_val == tc:
                keep_map[fn].append(r)

    created = []
    skipped = []

    for _, _, fn in targets_norm:
        rows = keep_map.get(fn, [])
        out_path = os.path.join(out_dir, fn)

        if not rows:
            skipped.append(fn)
            continue

        new_wb = Workbook()
        # Удалим дефолтный лист и создадим с нужным именем
        default_ws = new_wb.active
        new_wb.remove(default_ws)
        new_ws = new_wb.create_sheet(title=sheet_name)

        copy_sheet_filtered(ws, new_ws, rows)
        new_wb.save(out_path)
        created.append((fn, len(rows)))

    return out_dir, created, skipped


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Excel-разделитель по истцу и суду")
        self.geometry("740x330")
        self.resizable(False, False)

        self.input_path = None
        self.wb = None

        self._build_ui()

    def _build_ui(self):
        pad = 10

        frm = ttk.Frame(self, padding=pad)
        frm.pack(fill="both", expand=True)

        # Выбор файла
        row = 0
        ttk.Label(frm, text="1) Выберите Excel-файл:").grid(row=row, column=0, sticky="w")
        self.file_label = ttk.Label(frm, text="(файл не выбран)", width=70)
        self.file_label.grid(row=row, column=1, sticky="w", padx=(8, 0))

        ttk.Button(frm, text="Загрузить файл", command=self.pick_file).grid(row=row, column=2, padx=(8, 0))

        # Лист
        row += 1
        ttk.Label(frm, text="2) Лист:").grid(row=row, column=0, sticky="w", pady=(12, 0))
        self.sheet_var = tk.StringVar()
        self.sheet_cb = ttk.Combobox(frm, textvariable=self.sheet_var, state="disabled", width=45)
        self.sheet_cb.grid(row=row, column=1, sticky="w", pady=(12, 0))
        self.sheet_cb.bind("<<ComboboxSelected>>", lambda e: self.populate_columns())

        # Колонка истца
        row += 1
        ttk.Label(frm, text="3) Колонка 'Истец':").grid(row=row, column=0, sticky="w", pady=(12, 0))
        self.plaintiff_var = tk.StringVar()
        self.plaintiff_cb = ttk.Combobox(frm, textvariable=self.plaintiff_var, state="disabled", width=45)
        self.plaintiff_cb.grid(row=row, column=1, sticky="w", pady=(12, 0))

        # Колонка суда
        row += 1
        ttk.Label(frm, text="4) Колонка 'Суд':").grid(row=row, column=0, sticky="w", pady=(12, 0))
        self.court_var = tk.StringVar()
        self.court_cb = ttk.Combobox(frm, textvariable=self.court_var, state="disabled", width=45)
        self.court_cb.grid(row=row, column=1, sticky="w", pady=(12, 0))

        # Кнопка запуска
        row += 1
        self.run_btn = ttk.Button(frm, text="Разделить и сохранить 8 файлов", command=self.run, state="disabled")
        self.run_btn.grid(row=row, column=1, sticky="w", pady=(18, 0))

        # Подсказка
        row += 1
        hint = (
            "Файлы создаются в новой папке рядом с исходником.\n"
            "Если для какого-то истца/суда строк не найдено — файл не создаётся."
        )
        ttk.Label(frm, text=hint, foreground="#444").grid(row=row, column=1, sticky="w", pady=(12, 0))

    def pick_file(self):
        path = filedialog.askopenfilename(
            title="Выберите Excel-файл",
            filetypes=[("Excel files", "*.xlsx *.xlsm *.xltx *.xltm"), ("All files", "*.*")]
        )
        if not path:
            return
        self.input_path = path
        self.file_label.config(text=os.path.basename(path))

        try:
            self.wb = load_workbook(path, read_only=False)
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось открыть файл:\n{e}")
            self.wb = None
            return

        self.sheet_cb.config(state="readonly", values=self.wb.sheetnames)
        self.sheet_var.set(self.wb.sheetnames[0])
        self.populate_columns()

    def populate_columns(self):
        if not self.wb:
            return
        sheet = self.sheet_var.get()
        ws = self.wb[sheet]
        headers, _ = find_header_column_indices(ws)
        headers = [h for h in headers if h.strip() != ""]
        if not headers:
            messagebox.showerror("Ошибка", "Не найдены заголовки в первой строке листа.")
            return

        self.plaintiff_cb.config(state="readonly", values=headers)
        self.court_cb.config(state="readonly", values=headers)

        # Попытаемся авто-угадать
        def guess(headers_list, keywords):
            for h in headers_list:
                hn = norm_text(h)
                if any(k in hn for k in keywords):
                    return h
            return headers_list[0]

        self.plaintiff_var.set(guess(headers, ["истец", "взыскатель", "заявитель"]))
        self.court_var.set(guess(headers, ["суд", "судеб", "орган"]))

        self.run_btn.config(state="normal")

    def run(self):
        if not self.input_path:
            return

        sheet = self.sheet_var.get().strip()
        plaintiff_col = self.plaintiff_var.get().strip()
        court_col = self.court_var.get().strip()

        if not sheet or not plaintiff_col or not court_col:
            messagebox.showwarning("Внимание", "Выберите лист и обе колонки.")
            return

        try:
            out_dir, created, skipped = split_excel(self.input_path, sheet, plaintiff_col, court_col)
        except Exception as e:
            messagebox.showerror("Ошибка", str(e))
            return

        msg_lines = [f"Готово! Папка: {out_dir}", ""]
        if created:
            msg_lines.append("Создано:")
            for fn, n in created:
                msg_lines.append(f" • {fn}  (строк: {n})")
        else:
            msg_lines.append("Ничего не создано — не найдено совпадений по истцу/суду.")

        if skipped:
            msg_lines.append("")
            msg_lines.append("Пропущено (строк не найдено):")
            for fn in sorted(set(skipped)):
                msg_lines.append(f" • {fn}")

        messagebox.showinfo("Результат", "\n".join(msg_lines))


if __name__ == "__main__":
    App().mainloop()
